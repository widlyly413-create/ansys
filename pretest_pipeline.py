"""
预测试验证流水线（独立于主流程）

功能：
  在每个受试者正式测试前，先进行预测试（最大收缩 + 一次装饰实验），
  快速验证 sEMG 数据的合理性，决定是否继续后续测试。

使用方式：
  python pretest_pipeline.py                    # 处理所有受试者的预测试
  python pretest_pipeline.py 03zhangyining       # 只处理指定受试者

输出路径（独立于主流程）：
  pretest_results/<受试者>/
      ├── 预测试验证报告.xlsx       — MVC 值 + %MVC 统计数据
      └── 预测试验证图表.png        — 信号概览 + %MVC 柱状图
"""

import numpy as np
import os
import csv
import sys
from scipy.signal import butter, filtfilt, iirnotch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
matplotlib.rcParams['axes.unicode_minus'] = False


# ========== 常量 ==========

DATA_RAW_ROOT = r"e:\Coding_projects\SEMG_ANSYS\data_raw"
PROCESSED_ROOT = r"e:\Coding_projects\SEMG_ANSYS\data_processed"
PRETEST_RESULTS_ROOT = r"e:\Coding_projects\SEMG_ANSYS\pretest_results"

FS = 1000                    # 采样率 (Hz)
T_START = 10.0               # %MVC 分析起始时间 (秒)
T_END = 160.0                # %MVC 分析结束时间 (秒)
WINDOW_DURATION = 1.0        # 滑动窗口时长 (秒)
WINDOW_SIZE = int(WINDOW_DURATION * FS)

# 预测试子目录名
PRETEST_DIR = "预测试"

# 通道-肌肉映射
CHANNEL_MUSCLE_MAP = {
    1: "三角肌前束",
    2: "胸锁乳突肌",
    3: "斜方肌",
    4: "竖脊肌",
}

CHANNEL_COLS = [
    "Channel 1_三角肌前束",
    "Channel 2_胸锁乳突肌",
    "Channel 3_斜方肌",
    "Channel 4_竖脊肌",
]

MUSCLE_COLORS = {
    "Channel 1_三角肌前束": "#E74C3C",
    "Channel 2_胸锁乳突肌": "#2ECC71",
    "Channel 3_斜方肌": "#3498DB",
    "Channel 4_竖脊肌": "#F39C12",
}


# ========== 预处理函数（与 batch_preprocess.py 保持一致） ==========

def butter_bandpass_filter(data, lowcut=20, highcut=450, fs=1000, order=4):
    nyq = 0.5 * fs
    low = lowcut / nyq
    high = highcut / nyq
    b, a = butter(order, [low, high], btype='band')
    return filtfilt(b, a, data)


def notch_filter(data, cutoff=50, Q=30, fs=1000):
    nyq = 0.5 * fs
    w0 = cutoff / nyq
    b, a = iirnotch(w0, Q)
    return filtfilt(b, a, data)


def preprocess_semg(raw_data, fs=1000):
    step1 = raw_data - np.mean(raw_data)
    step2 = butter_bandpass_filter(step1, lowcut=20, highcut=450, fs=fs, order=4)
    step3 = notch_filter(step2, cutoff=50, Q=30, fs=fs)
    return np.abs(step3)


# ========== 文件读写函数 ==========

def read_csv_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return None, None
    return rows[0], rows[1:]


def write_csv_file(filepath, header, data_rows):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(data_rows)


def get_column_index(header, col_name):
    for i, h in enumerate(header):
        if h.strip() == col_name:
            return i
    return -1


def extract_column(data_rows, col_idx):
    return np.array([float(row[col_idx]) for row in data_rows], dtype=np.float64)


# ========== 分析和统计函数 ==========

def extract_mvc_max(rectified_signal, fs=1000, window_duration=1.0):
    """提取单通道 MVC 极大值（1秒滑动窗口均值）"""
    window_size = int(window_duration * fs)
    if len(rectified_signal) < window_size:
        return None
    window = np.ones(window_size) / window_size
    smoothed = np.convolve(rectified_signal, window, mode='valid')
    return float(np.max(smoothed))


def sliding_rms(signal, window_size):
    """计算滑动均方根值"""
    squared = signal ** 2
    window = np.ones(window_size) / window_size
    mean_squared = np.convolve(squared, window, mode='valid')
    return np.sqrt(mean_squared)


def compute_percent_mvc_stats(rms_seq, mvc_value):
    """计算 %MVC 统计指标"""
    if mvc_value <= 0:
        return (0.0, 0.0, 0.0, 0.0)
    percent_mvc = (rms_seq / mvc_value) * 100.0
    return (
        float(np.mean(percent_mvc)),
        float(np.percentile(percent_mvc, 10)),
        float(np.percentile(percent_mvc, 50)),
        float(np.percentile(percent_mvc, 90)),
    )


# ========== 数据预处理 ==========

def preprocess_pretest_files(subject_dir):
    """
    对指定受试者的预测试文件夹进行预处理。
    返回: (mvc_file_processed, decor_file_processed) 预处理后的文件路径，或 None
    """
    raw_pretest_path = os.path.join(DATA_RAW_ROOT, subject_dir, PRETEST_DIR)
    if not os.path.isdir(raw_pretest_path):
        return None, None

    out_pretest_path = os.path.join(PROCESSED_ROOT, subject_dir, PRETEST_DIR)
    os.makedirs(out_pretest_path, exist_ok=True)

    found_mvc = None
    found_decor = None

    for fname in os.listdir(raw_pretest_path):
        fpath = os.path.join(raw_pretest_path, fname)
        if not os.path.isfile(fpath):
            continue

        out_path = os.path.join(out_pretest_path, fname)

        # 如果已预处理过，直接使用
        if os.path.exists(out_path):
            print(f"  ⏭ 已预处理: {fname}")
        else:
            print(f"  预处理: {fname}")
            header, data_rows = read_csv_file(fpath)
            if header is None:
                print(f"    文件为空，跳过")
                continue

            required_cols = [f"Channel {i}" for i in range(1, 5)]
            col_indices = {}
            for col_name in required_cols:
                idx = get_column_index(header, col_name)
                if idx == -1:
                    break
                col_indices[col_name] = idx
            if len(col_indices) < 4:
                print(f"    缺少通道列，跳过")
                continue

            processed_cols = {}
            for ch_idx in range(1, 5):
                col_name = f"Channel {ch_idx}"
                raw_signal = extract_column(data_rows, col_indices[col_name])
                processed_signal = preprocess_semg(raw_signal, fs=1000)
                muscle_name = CHANNEL_MUSCLE_MAP[ch_idx]
                new_col_name = f"{col_name}_{muscle_name}"
                processed_cols[new_col_name] = processed_signal

            output_header = ["timestamp"]
            for ch_idx in range(1, 5):
                output_header.append(f"Channel {ch_idx}_{CHANNEL_MUSCLE_MAP[ch_idx]}")

            ts_idx = get_column_index(header, "timestamp")
            output_rows = []
            n_samples = len(data_rows)
            for i in range(n_samples):
                row = []
                if ts_idx != -1:
                    row.append(data_rows[i][ts_idx])
                for ch_idx in range(1, 5):
                    new_col_name = f"Channel {ch_idx}_{CHANNEL_MUSCLE_MAP[ch_idx]}"
                    row.append(f"{processed_cols[new_col_name][i]:.6f}")
                output_rows.append(row)

            write_csv_file(out_path, output_header, output_rows)
            print(f"    已保存 -> {out_path}")

        # 分类：最大收缩 vs 装饰
        if "最大收缩" in fname:
            found_mvc = out_path
        elif "装饰" in fname:
            found_decor = out_path

    return found_mvc, found_decor


# ========== 生成验证报告 ==========

def generate_report(subject_dir, mvc_file, decor_file):
    """生成 MVC 值和 %MVC 统计的 Excel 报告"""
    print(f"\n  生成验证报告...")
    report_dir = os.path.join(PRETEST_RESULTS_ROOT, subject_dir)
    os.makedirs(report_dir, exist_ok=True)
    report_path = os.path.join(report_dir, "预测试验证报告.xlsx")

    wb = Workbook()

    # --- Sheet 1: MVC 值 ---
    ws_mvc = wb.active
    ws_mvc.title = "MVC 值"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 读取 MVC 文件
    mvc_header, mvc_data = read_csv_file(mvc_file)
    if mvc_header is None:
        print("  错误：无法读取 MVC 预处理文件")
        return None

    mvc_values = {}
    for col_name in CHANNEL_COLS:
        col_idx = get_column_index(mvc_header, col_name)
        if col_idx == -1:
            mvc_values[col_name] = None
            continue
        signal = extract_column(mvc_data, col_idx)
        mvc_val = extract_mvc_max(signal, fs=FS)
        mvc_values[col_name] = mvc_val

    mvc_headers = ["通道", "肌肉", "MVC_极大值"]
    for col_idx, h in enumerate(mvc_headers, 1):
        cell = ws_mvc.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for r_idx, col_name in enumerate(CHANNEL_COLS, 2):
        ws_mvc.cell(row=r_idx, column=1, value=col_name).border = thin_border
        muscle_name = col_name.split("_", 1)[1]
        ws_mvc.cell(row=r_idx, column=2, value=muscle_name).border = thin_border
        val = mvc_values.get(col_name, None)
        ws_mvc.cell(row=r_idx, column=3, value=round(val, 4) if val else "N/A").border = thin_border

    # 数据合理性判断
    ws_mvc.cell(row=7, column=1, value="数据合理性检查：").font = Font(bold=True, color="CC0000")
    ws_mvc.cell(row=8, column=1, value="MVC 值过低（<10）可能表示信号异常")
    ws_mvc.cell(row=9, column=1, value="MVC 值过高（>1000）可能表示噪声干扰")
    for r_idx, col_name in enumerate(CHANNEL_COLS, 2):
        val = mvc_values.get(col_name, None)
        if val is not None:
            status = "✅ 正常" if 10 <= val <= 1000 else "⚠️ 异常"
            ws_mvc.cell(row=r_idx, column=4, value=status).border = thin_border

    for col in range(1, 5):
        ws_mvc.column_dimensions[chr(64 + col)].width = 22

    # --- Sheet 2: %MVC 统计 ---
    ws_pct = wb.create_sheet(title="%MVC 统计")
    decor_name = os.path.basename(decor_file) if decor_file else "装饰任务"

    pct_headers = ["通道", "肌肉", "Mean_%MVC", "APDF_10%", "APDF_50%", "APDF_90%"]
    for col_idx, h in enumerate(pct_headers, 1):
        cell = ws_pct.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ws_pct.cell(row=1, column=7, value=f"任务: {decor_name}")
    ws_pct.cell(row=1, column=7).font = Font(bold=True, color="333399")

    if decor_file and mvc_values:
        decor_header, decor_data = read_csv_file(decor_file)
        if decor_header:
            for r_idx, col_name in enumerate(CHANNEL_COLS, 2):
                mvc_val = mvc_values.get(col_name, None)
                if mvc_val is None or mvc_val <= 0:
                    continue

                col_idx = get_column_index(decor_header, col_name)
                if col_idx == -1:
                    continue
                signal = extract_column(decor_data, col_idx)

                # 截取分析区间
                start_idx = int(T_START * FS)
                end_idx = int(T_END * FS)
                if end_idx > len(signal):
                    end_idx = len(signal)
                signal_seg = signal[start_idx:end_idx]

                if len(signal_seg) < WINDOW_SIZE:
                    continue

                rms_seq = sliding_rms(signal_seg, WINDOW_SIZE)
                mean_val, p10, p50, p90 = compute_percent_mvc_stats(rms_seq, mvc_val)

                muscle_name = col_name.split("_", 1)[1]
                ws_pct.cell(row=r_idx, column=1, value=col_name).border = thin_border
                ws_pct.cell(row=r_idx, column=2, value=muscle_name).border = thin_border
                ws_pct.cell(row=r_idx, column=3, value=round(mean_val, 4)).border = thin_border
                ws_pct.cell(row=r_idx, column=4, value=round(p10, 4)).border = thin_border
                ws_pct.cell(row=r_idx, column=5, value=round(p50, 4)).border = thin_border
                ws_pct.cell(row=r_idx, column=6, value=round(p90, 4)).border = thin_border

                # 合理性检查
                status = "✅ 合理" if mean_val < 80 else "⚠️ %MVC偏高，检查MVC值"
                ws_pct.cell(row=r_idx, column=7, value=status).border = thin_border

        for col in range(1, 8):
            ws_pct.column_dimensions[chr(64 + col)].width = 20

    wb.save(report_path)
    print(f"  ✅ 报告已保存: {report_path}")
    return report_path


# ========== 生成验证图表 ==========

def generate_chart(subject_dir, mvc_file, decor_file, report_path):
    """生成预测试验证图表"""
    print(f"  生成验证图表...")

    # 读取数据
    mvc_header, mvc_data = read_csv_file(mvc_file)
    decor_header, decor_data = read_csv_file(decor_file) if decor_file else (None, None)

    if mvc_header is None:
        print("  错误：无法读取数据，跳过图表")
        return

    # 提取 MVC 值
    mvc_values = {}
    for col_name in CHANNEL_COLS:
        col_idx = get_column_index(mvc_header, col_name)
        if col_idx == -1:
            continue
        signal = extract_column(mvc_data, col_idx)
        mvc_val = extract_mvc_max(signal, fs=FS)
        mvc_values[col_name] = mvc_val

    # 提取装饰任务的 %MVC 统计值
    pct_stats = {}
    if decor_header and decor_data and mvc_values:
        for col_name in CHANNEL_COLS:
            mvc_val = mvc_values.get(col_name, None)
            if mvc_val is None or mvc_val <= 0:
                continue
            col_idx = get_column_index(decor_header, col_name)
            if col_idx == -1:
                continue
            signal = extract_column(decor_data, col_idx)
            start_idx = int(T_START * FS)
            end_idx = int(T_END * FS)
            if end_idx > len(signal):
                end_idx = len(signal)
            signal_seg = signal[start_idx:end_idx]
            if len(signal_seg) < WINDOW_SIZE:
                continue
            rms_seq = sliding_rms(signal_seg, WINDOW_SIZE)
            mean_val, p10, p50, p90 = compute_percent_mvc_stats(rms_seq, mvc_val)
            pct_stats[col_name] = {"Mean": mean_val, "P10": p10, "P50": p50, "P90": p90}

    # 开始绘图
    fig = plt.figure(figsize=(18, 10))
    gs = fig.add_gridspec(2, 3, hspace=0.35, wspace=0.30)

    # ---- 图1: MVC 柱状图 ----
    ax1 = fig.add_subplot(gs[0, 0])
    muscles_short = [c.split("_", 1)[1] for c in CHANNEL_COLS]
    mvc_vals = [mvc_values.get(c, 0) for c in CHANNEL_COLS]
    colors = [MUSCLE_COLORS[c] for c in CHANNEL_COLS]
    bars = ax1.bar(muscles_short, mvc_vals, color=colors, edgecolor='white', linewidth=0.5)
    for bar, val in zip(bars, mvc_vals):
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(mvc_vals)*0.02,
                 f'{val:.1f}', ha='center', va='bottom', fontsize=9)
    ax1.set_title("预测试 - MVC 最大值", fontsize=14, fontweight='bold')
    ax1.set_ylabel("MVC (整流后 μV)")
    ax1.grid(axis='y', alpha=0.3)
    ax1.set_ylim(0, max(mvc_vals) * 1.25 if max(mvc_vals) > 0 else 100)

    # ---- 图2: 装饰任务 %MVC 对比 ----
    ax2 = fig.add_subplot(gs[0, 1])
    if pct_stats:
        x = np.arange(len(CHANNEL_COLS))
        bar_width = 0.5
        means = [pct_stats[c]["Mean"] for c in CHANNEL_COLS]
        p10s = [pct_stats[c]["P10"] for c in CHANNEL_COLS]
        p50s = [pct_stats[c]["P50"] for c in CHANNEL_COLS]
        p90s = [pct_stats[c]["P90"] for c in CHANNEL_COLS]

        bars_mean = ax2.bar(x, means, bar_width, color=colors, edgecolor='white', linewidth=0.5)
        for bar, val in zip(bars_mean, means):
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(means)*0.02,
                     f'{val:.1f}%', ha='center', va='bottom', fontsize=8)
        ax2.set_xticks(x)
        ax2.set_xticklabels(muscles_short)
        ax2.set_title("装饰任务 - Mean %MVC", fontsize=14, fontweight='bold')
        ax2.set_ylabel("%MVC")
        ax2.grid(axis='y', alpha=0.3)
        ax2.set_ylim(0, max(means) * 1.4 if max(means) > 0 else 20)

        # 同时在右侧用表格显示 APDF 值
        table_data = []
        for i, c in enumerate(CHANNEL_COLS):
            table_data.append([muscles_short[i], f'{p10s[i]:.1f}', f'{p50s[i]:.1f}', f'{p90s[i]:.1f}'])
        col_labels = ['肌肉', 'P10', 'P50', 'P90']
        table_ax = fig.add_subplot(gs[0, 2])
        table_ax.axis('off')
        table = table_ax.table(cellText=table_data, colLabels=col_labels,
                               cellLoc='center', loc='center')
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1, 1.8)
        table_ax.set_title("APDF 百分位值 (%MVC)", fontsize=14, fontweight='bold', pad=20)
    else:
        ax2.text(0.5, 0.5, "无装饰任务数据", ha='center', va='center', fontsize=14)
        ax2.set_title("装饰任务 - Mean %MVC", fontsize=14, fontweight='bold')

    # ---- 图3: 信号波形概览（装饰任务的前 5 秒预览） ----
    ax3 = fig.add_subplot(gs[1, :])
    if decor_header and decor_data:
        time_range = min(5000, len(decor_data))  # 取前 5 秒
        t = np.arange(time_range) / FS
        for i, col_name in enumerate(CHANNEL_COLS):
            col_idx = get_column_index(decor_header, col_name)
            if col_idx == -1:
                continue
            signal = extract_column(decor_data, col_idx)[:time_range]
            offset = i * 200  # 上下偏移便于显示
            ax3.plot(t, signal + offset, color=MUSCLE_COLORS[col_name],
                     label=col_name.split("_", 1)[1], linewidth=0.6)
        ax3.set_xlabel("时间 (秒)")
        ax3.set_ylabel("信号幅值 (偏移显示)")
        ax3.set_title("装饰任务 - 预处理后信号预览（前 5 秒）", fontsize=14, fontweight='bold')
        ax3.legend(fontsize=9, loc='upper right')
        ax3.grid(alpha=0.3)
        ax3.set_yticks([])  # 隐藏 y 轴刻度
    else:
        ax3.text(0.5, 0.5, "无装饰任务数据", ha='center', va='center', fontsize=14)
        ax3.set_title("装饰任务 - 信号预览", fontsize=14, fontweight='bold')

    fig.suptitle(f"受试者: {subject_dir} — 预测试数据验证",
                 fontsize=18, fontweight='bold', y=1.02)
    plt.tight_layout()

    chart_path = os.path.join(PRETEST_RESULTS_ROOT, subject_dir, "预测试验证图表.png")
    plt.savefig(chart_path, dpi=200, bbox_inches='tight')
    plt.close()
    print(f"  ✅ 图表已保存: {chart_path}")

    return chart_path


# ========== 打印终端摘要 ==========

def print_summary(subject_dir, mvc_values, pct_stats, chart_path, report_path):
    """在终端打印预测试结果摘要"""
    print(f"\n{'='*60}")
    print(f"  受试者: {subject_dir} — 预测试验证结果")
    print(f"{'='*60}")

    print(f"\n  📊 MVC 极大值：")
    print(f"  {'通道':<30} {'MVC值':<12} {'状态':<10}")
    print(f"  {'-'*54}")
    for col_name in CHANNEL_COLS:
        val = mvc_values.get(col_name, None)
        muscle = col_name.split("_", 1)[1]
        if val is not None:
            status = "✅" if 10 <= val <= 1000 else "⚠️"
            print(f"  {muscle:<30} {val:<12.4f} {status:<10}")
        else:
            print(f"  {muscle:<30} {'N/A':<12} {'❌':<10}")

    if pct_stats:
        print(f"\n  📈 装饰任务 %MVC（Mean）：")
        print(f"  {'肌肉':<30} {'Mean%MVC':<12}")
        print(f"  {'-'*44}")
        for col_name in CHANNEL_COLS:
            if col_name in pct_stats:
                muscle = col_name.split("_", 1)[1]
                mean_val = pct_stats[col_name]["Mean"]
                status = "✅" if mean_val < 80 else "⚠️"
                print(f"  {muscle:<30} {mean_val:<12.4f} {status}")

    print(f"\n  📁 输出文件：")
    if report_path:
        print(f"     报告: {report_path}")
    if chart_path:
        print(f"     图表: {chart_path}")
    print(f"{'='*60}\n")


# ========== 主程序 ==========

def main():
    target_subject = sys.argv[1] if len(sys.argv) > 1 else None

    os.makedirs(PRETEST_RESULTS_ROOT, exist_ok=True)

    print("=" * 60)
    print("  sEMG 预测试验证流水线")
    print("  用于在正式测试前快速验证数据合理性")
    print("=" * 60)

    if target_subject:
        subject_dirs = [d for d in os.listdir(DATA_RAW_ROOT)
                        if os.path.isdir(os.path.join(DATA_RAW_ROOT, d)) and d == target_subject]
        if not subject_dirs:
            print(f"错误：未找到受试者目录 '{target_subject}'")
            sys.exit(1)
    else:
        subject_dirs = sorted([d for d in os.listdir(DATA_RAW_ROOT)
                               if os.path.isdir(os.path.join(DATA_RAW_ROOT, d))])

    processed_any = False

    for subject_dir in subject_dirs:
        print(f"\n{'='*60}")
        print(f"  受试者: {subject_dir}")
        print(f"{'='*60}")

        # 1. 预处理预测试文件
        mvc_file, decor_file = preprocess_pretest_files(subject_dir)
        if mvc_file is None:
            print(f"  ⚠ 未找到预测试目录或最大收缩文件，跳过")
            print(f"  请在 data_raw/{subject_dir}/预测试/ 下放置预测试文件")
            continue

        if decor_file is None:
            print(f"  ⚠ 未找到预测试装饰任务文件，跳过")
            print(f"  请在 data_raw/{subject_dir}/预测试/ 下放置装饰实验文件")
            continue

        processed_any = True

        # 2. 提取 MVC 值
        mvc_header, mvc_data = read_csv_file(mvc_file)
        mvc_values = {}
        for col_name in CHANNEL_COLS:
            col_idx = get_column_index(mvc_header, col_name)
            if col_idx == -1:
                continue
            signal = extract_column(mvc_data, col_idx)
            mvc_val = extract_mvc_max(signal, fs=FS)
            mvc_values[col_name] = mvc_val

        # 3. 计算装饰任务的 %MVC 统计
        pct_stats = {}
        decor_header, decor_data = read_csv_file(decor_file)
        if decor_header and mvc_values:
            for col_name in CHANNEL_COLS:
                mvc_val = mvc_values.get(col_name, None)
                if mvc_val is None or mvc_val <= 0:
                    continue
                col_idx = get_column_index(decor_header, col_name)
                if col_idx == -1:
                    continue
                signal = extract_column(decor_data, col_idx)
                start_idx = int(T_START * FS)
                end_idx = int(T_END * FS)
                if end_idx > len(signal):
                    end_idx = len(signal)
                signal_seg = signal[start_idx:end_idx]
                if len(signal_seg) < WINDOW_SIZE:
                    continue
                rms_seq = sliding_rms(signal_seg, WINDOW_SIZE)
                mean_val, p10, p50, p90 = compute_percent_mvc_stats(rms_seq, mvc_val)
                pct_stats[col_name] = {"Mean": mean_val, "P10": p10, "P50": p50, "P90": p90}

        # 4. 生成报告和图表
        report_path = generate_report(subject_dir, mvc_file, decor_file)
        chart_path = generate_chart(subject_dir, mvc_file, decor_file, report_path)

        # 5. 打印摘要
        print_summary(subject_dir, mvc_values, pct_stats, chart_path, report_path)

    if not processed_any:
        print(f"\n⚠ 没有处理任何预测试数据。")
        print(f"使用说明：")
        print(f"  1. 在 data_raw/<受试者>/预测试/ 下放入预测试文件")
        print(f"  2. 文件名需包含 '最大收缩' 和 '装饰' 关键字")
        print(f"  3. 运行: python pretest_pipeline.py")
    else:
        print(f"\n✅ 预测试验证完成！请查看 pretest_results/ 目录下的报告和图表。")
        print(f"   - 验证图表的 %MVC 值是否合理（一般 < 30% 为正常）")
        print(f"   - 检查 MVC 值是否在合理范围内（10~1000）")
        print(f"   - 确认信号预览无明显噪声干扰")
        print(f"\n  数据合理 → 继续正式测试")
        print(f"  数据异常 → 检查设备或重新测试")


if __name__ == "__main__":
    main()