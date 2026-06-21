import matplotlib.pyplot as plt
import matplotlib
matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
matplotlib.rcParams['axes.unicode_minus'] = False

import numpy as np
import os
from openpyxl import load_workbook

# ========== 路径常量 ==========

EXCEL_PATH = r"e:\Coding_projects\SEMG_ANSYS\analysis_results.xlsx"
SUBJECT_INFO_PATH = r"e:\Coding_projects\SEMG_ANSYS\受试者信息收集表.xlsx"
OUTPUT_DIR = r"e:\Coding_projects\SEMG_ANSYS\figures"

# 肌肉名称简写（用于图表 x 轴）
MUSCLE_SHORT = {
    "Channel 1_三角肌前束": "三角肌前束",
    "Channel 2_胸锁乳突肌": "胸锁乳突肌",
    "Channel 3_斜方肌": "斜方肌",
    "Channel 4_竖脊肌": "竖脊肌",
}

MUSCLE_COLORS = {
    "Channel 1_三角肌前束": "#E74C3C",
    "Channel 2_胸锁乳突肌": "#2ECC71",
    "Channel 3_斜方肌": "#3498DB",
    "Channel 4_竖脊肌": "#F39C12",
}

# 指标名称（用于子图标题）
METRICS = ["Mean_%MVC", "APDF_10%", "APDF_50%", "APDF_90%"]
METRIC_LABELS = {
    "Mean_%MVC": "Mean %MVC\n(平均负荷)",
    "APDF_10%": "APDF 10%\n(静态负荷)",
    "APDF_50%": "APDF 50%\n(中值负荷)",
    "APDF_90%": "APDF 90%\n(峰值负荷)",
}


# ========== 数据加载 ==========

def load_data(excel_path):
    """从 Excel 读取所有数据，返回结构化字典"""
    wb = load_workbook(excel_path, data_only=True)

    data = {}  # data[subject][task_name][channel] = {metric: value, ...}

    for sheet_name in wb.sheetnames:
        if sheet_name == "汇总":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        subject_data = {}
        for row in rows:
            task = row[0]
            channel = row[1]
            metrics = {
                "Mean_%MVC": row[2],
                "APDF_10%": row[3],
                "APDF_50%": row[4],
                "APDF_90%": row[5],
            }
            if task not in subject_data:
                subject_data[task] = {}
            subject_data[task][channel] = metrics

        data[sheet_name] = subject_data

    return data


def load_subject_info(info_path):
    """读取受试者信息收集表，返回 subject_id → {desk_diff_装饰, desk_diff_打磨}"""
    wb = load_workbook(info_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    info = {}
    # Row 0: Subject_ID, subj1, subj2, ...
    subject_ids = [str(v) for v in rows[0][1:] if v is not None]

    # 建立 row_label → col_index 映射
    label_to_col = {}
    for col_idx, val in enumerate(rows[0]):
        if val is not None:
            label_to_col["Subject_ID"] = 0  # placeholder
            break

    # 遍历各行找需要的字段
    desk_diff_keys = {"装饰桌高差值", "打磨桌高差值"}
    found = {}

    for row in rows:
        label = str(row[0]).strip() if row[0] is not None else ""
        if label in desk_diff_keys:
            found[label] = {}
            for i, subj_id in enumerate(subject_ids):
                val = row[i + 1]
                found[label][subj_id] = float(val) if val is not None else None

    # 重新组织 info
    for subj_id in subject_ids:
        info[subj_id] = {
            "装饰桌高差": found.get("装饰桌高差值", {}).get(subj_id, None),
            "打磨桌高差": found.get("打磨桌高差值", {}).get(subj_id, None),
        }

    wb.close()
    return info


# ========== 绘图函数 ==========

def plot_comparison(data, subject, operation_type, channels, output_dir):
    """
    为某个受试者的某类操作绘制对比图
    operation_type: '装饰' 或 '打磨'
    """
    # 找到对应的传统/优化任务名
    trad_task = None
    opt_task = None
    for task in data[subject]:
        if operation_type in task and "传统" in task:
            trad_task = task
        if operation_type in task and "优化" in task:
            opt_task = task

    if trad_task is None or opt_task is None:
        print(f"  {subject} 缺少 {operation_type} 的 传统/优化 数据，跳过")
        return

    # 构建图表
    n_metrics = len(METRICS)
    n_muscles = len(channels)

    fig, axes = plt.subplots(1, n_metrics, figsize=(5.5 * n_metrics, 5))
    if n_metrics == 1:
        axes = [axes]

    x = np.arange(n_muscles)       # 肌肉位置
    bar_width = 0.35
    colors = {"传统": "#4A7FB5", "优化": "#E8834A"}

    for m_idx, metric in enumerate(METRICS):
        ax = axes[m_idx]

        trad_values = []
        opt_values = []
        for ch in channels:
            trad_values.append(data[subject][trad_task][ch][metric])
            opt_values.append(data[subject][opt_task][ch][metric])

        # 绘制分组柱状图
        bars1 = ax.bar(x - bar_width / 2, trad_values, bar_width,
                       label=f"{operation_type}传统", color=colors["传统"],
                       edgecolor='white', linewidth=0.5)
        bars2 = ax.bar(x + bar_width / 2, opt_values, bar_width,
                       label=f"{operation_type}优化", color=colors["优化"],
                       edgecolor='white', linewidth=0.5)

        # 在柱子上标数值
        for bar in bars1:
            h = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, h + 0.5,
                    f'{h:.1f}', ha='center', va='bottom', fontsize=8)
        for bar in bars2:
            h = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, h + 0.5,
                    f'{h:.1f}', ha='center', va='bottom', fontsize=8)

        ax.set_xticks(x)
        ax.set_xticklabels([MUSCLE_SHORT[ch] for ch in channels], fontsize=10)
        ax.set_title(METRIC_LABELS[metric], fontsize=13, fontweight='bold', pad=10)
        ax.set_ylabel("%MVC", fontsize=11)
        ax.legend(fontsize=9)
        ax.grid(axis='y', alpha=0.3)
        # 确保 y 轴从 0 开始，留 10% 余量
        y_max = max(max(trad_values), max(opt_values)) * 1.2
        ax.set_ylim(0, max(y_max, 10))

    fig.suptitle(f"{subject} — {operation_type}工序：传统 vs 优化",
                 fontsize=16, fontweight='bold', y=1.02)
    plt.tight_layout()

    # 保存
    os.makedirs(output_dir, exist_ok=True)
    safe_subject = subject.replace(" ", "_")
    save_path = os.path.join(output_dir, f"{safe_subject}_{operation_type}_对比.png")
    plt.savefig(save_path, dpi=200, bbox_inches='tight')
    plt.close()
    print(f"  已保存: {save_path}")


def plot_diff_vs_desk_height(data, subject_info, operation_type, channels, output_dir):
    """
    绘制优化-传统 %MVC 差值与桌高差值的关系散点图
    每个指标一个子图，4 块肌肉用不同颜色线条连接 2 个受试者的数据点
    """
    subjects = sorted([s for s in data.keys() if s != "汇总"])

    # 收集每个受试者的桌高差和 %MVC 差值
    desk_diff_key = f"{operation_type}桌高差"
    trad_suffix, opt_suffix = "传统", "优化"

    n_metrics = len(METRICS)
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    axes = axes.flatten()

    for m_idx, metric in enumerate(METRICS):
        ax = axes[m_idx]

        for ch in channels:
            x_vals = []
            y_vals = []
            labels = []

            for subj in subjects:
                info = subject_info.get(subj, {})
                desk_diff = info.get(desk_diff_key)
                if desk_diff is None:
                    continue

                trad_task = None
                opt_task = None
                for task in data.get(subj, {}):
                    if operation_type in task and trad_suffix in task:
                        trad_task = task
                    if operation_type in task and opt_suffix in task:
                        opt_task = task

                if trad_task is None or opt_task is None:
                    continue
                if ch not in data[subj][trad_task] or ch not in data[subj][opt_task]:
                    continue

                trad_val = data[subj][trad_task][ch][metric]
                opt_val = data[subj][opt_task][ch][metric]
                diff = opt_val - trad_val

                x_vals.append(desk_diff)
                y_vals.append(diff)
                labels.append(subj)

            if len(x_vals) >= 2:
                # 画连线 + 散点
                ax.plot(x_vals, y_vals, '-o', color=MUSCLE_COLORS[ch],
                        linewidth=1, markersize=4, label=MUSCLE_SHORT[ch])
                # 标注差值（贴着数据点上边缘）
                for x, y, label in zip(x_vals, y_vals, labels):
                    ax.annotate(f"{y:.1f}", (x, y),
                                textcoords="offset points", xytext=(0, 0),
                                ha='center', va='bottom', fontsize=9,
                                color=MUSCLE_COLORS[ch])
            elif len(x_vals) == 1:
                ax.scatter(x_vals, y_vals, color=MUSCLE_COLORS[ch],
                           s=40, label=MUSCLE_SHORT[ch], zorder=5)

        ax.axhline(y=0, color='gray', linestyle='--', linewidth=0.8, alpha=0.6)
        ax.axvline(x=0, color='gray', linestyle='--', linewidth=0.8, alpha=0.6)
        ax.set_xlabel(f"{operation_type}桌高差值 (mm)", fontsize=11)
        ax.set_ylabel("%MVC 差值 (优化 - 传统)", fontsize=11)
        ax.set_title(METRIC_LABELS[metric], fontsize=13, fontweight='bold')
        ax.legend(fontsize=9, loc='best')
        ax.grid(True, alpha=0.3)

    fig.suptitle(f"{operation_type}工序：%MVC 差值 vs 桌高差值",
                 fontsize=16, fontweight='bold', y=1.01)
    plt.tight_layout()

    os.makedirs(output_dir, exist_ok=True)
    save_path = os.path.join(output_dir, f"差值_vs_桌高差_{operation_type}.png")
    plt.savefig(save_path, dpi=200, bbox_inches='tight')
    plt.close()
    print(f"  已保存: {save_path}")


# ========== 主程序 ==========

def main():
    print("加载 Excel 数据...")
    data = load_data(EXCEL_PATH)
    subjects = sorted([s for s in data.keys() if s != "汇总"])
    print(f"  受试者: {subjects}")

    print("加载受试者信息收集表...")
    subject_info = load_subject_info(SUBJECT_INFO_PATH)
    for subj in subjects:
        info = subject_info.get(subj, {})
        print(f"  {subj}: 装饰桌高差={info.get('装饰桌高差', 'N/A')}mm, "
              f"打磨桌高差={info.get('打磨桌高差', 'N/A')}mm")

    channels = [
        "Channel 1_三角肌前束",
        "Channel 2_胸锁乳突肌",
        "Channel 3_斜方肌",
        "Channel 4_竖脊肌",
    ]

    operation_types = ["装饰", "打磨"]

    print("\n生成单个受试者对比图...")
    for subject in subjects:
        for op in operation_types:
            plot_comparison(data, subject, op, channels, OUTPUT_DIR)

    print("\n生成差值 vs 桌高差关系图...")
    for op in operation_types:
        plot_diff_vs_desk_height(data, subject_info, op, channels, OUTPUT_DIR)

    print(f"\n✅ 所有图表已保存至: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()