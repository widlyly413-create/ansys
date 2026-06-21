"""
合并 MVC 极大值脚本

用途：对比 zhangyining 的两个最大收缩测试文件，对每个通道取最大值，
      更新 MVC_values.csv，然后重新运行后续分析。
"""

import numpy as np
import os
import csv
import sys

# ========== 常量 ==========

PROCESSED_ROOT = r"e:\Coding_projects\SEMG_ANSYS\data_processed"
MVC_CSV_PATH = r"e:\Coding_projects\SEMG_ANSYS\MVC_values.csv"
OUTPUT_EXCEL = r"e:\Coding_projects\SEMG_ANSYS\analysis_results.xlsx"
FIGURES_DIR = r"e:\Coding_projects\SEMG_ANSYS\figures"

SUBJECT = "03zhangyining"
MVC_FILES = ["zhangyining最大收缩测试", "zhangyining_5最大收缩测试补充"]

CHANNEL_COLS = [
    "Channel 1_三角肌前束",
    "Channel 2_胸锁乳突肌",
    "Channel 3_斜方肌",
    "Channel 4_竖脊肌",
]


# ========== 工具函数 ==========

def read_csv_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return None, None
    return rows[0], rows[1:]


def get_column_index(header, col_name):
    for i, h in enumerate(header):
        if h.strip() == col_name:
            return i
    return -1


def extract_mvc_max(rectified_mvc_signal, fs=1000, window_duration=1.0):
    """提取单通道 MVC 极大值（1秒滑动窗口均值）"""
    window_size = int(window_duration * fs)
    if len(rectified_mvc_signal) < window_size:
        return None
    window = np.ones(window_size) / window_size
    smoothed_signal = np.convolve(rectified_mvc_signal, window, mode='valid')
    return np.max(smoothed_signal)


# ========== 主程序 ==========

def main():
    print("=" * 60)
    print("  MVC 对比合并：取两个最大收缩测试文件的最大值")
    print("=" * 60)

    subject_path = os.path.join(PROCESSED_ROOT, SUBJECT)
    if not os.path.isdir(subject_path):
        print(f"错误：找不到受试者目录 {subject_path}")
        sys.exit(1)

    # 1. 读取两个 MVC 文件的数据
    all_results = {}
    for mvc_file in MVC_FILES:
        file_path = os.path.join(subject_path, mvc_file)
        if not os.path.isfile(file_path):
            print(f"  ⚠ 文件不存在，跳过: {file_path}")
            continue

        print(f"\n--- 处理: {mvc_file} ---")
        header, data_rows = read_csv_file(file_path)
        if header is None:
            print(f"  文件为空，跳过")
            continue

        for col_name in CHANNEL_COLS:
            col_idx = get_column_index(header, col_name)
            if col_idx == -1:
                print(f"  列 {col_name} 不存在，跳过")
                continue

            signal = np.array(
                [float(row[col_idx]) for row in data_rows], dtype=np.float64
            )
            mvc_value = extract_mvc_max(signal, fs=1000, window_duration=1.0)
            if mvc_value is None:
                print(f"  通道 {col_name}: 信号太短，无法计算")
                continue

            print(f"  {col_name}: MVC = {mvc_value:.4f}")

            if col_name not in all_results:
                all_results[col_name] = []
            all_results[col_name].append((mvc_file, mvc_value))

    # 2. 对比取最大值
    print("\n" + "=" * 60)
    print("  MVC 对比结果（取最大值）")
    print("=" * 60)
    
    final_mvc = {}
    for col_name in CHANNEL_COLS:
        if col_name not in all_results or len(all_results[col_name]) == 0:
            print(f"  ⚠ {col_name}: 无有效数据")
            continue

        values = all_results[col_name]
        max_entry = max(values, key=lambda x: x[1])
        final_mvc[col_name] = max_entry[1]

        print(f"\n  {col_name}:")
        for file_name, val in values:
            indicator = " ← 取最大值" if val == max_entry[1] else ""
            print(f"    {file_name:<30}: {val:.4f}{indicator}")
        print(f"    {'最终取值':<30}: {max_entry[1]:.4f}")

    # 3. 读取现有的 MVC_values.csv
    print(f"\n  更新 {MVC_CSV_PATH} ...")
    header, data_rows = read_csv_file(MVC_CSV_PATH)
    if header is None:
        print("  错误：无法读取 MVC_values.csv")
        sys.exit(1)

    subj_idx = get_column_index(header, "受试者")
    chan_idx = get_column_index(header, "通道_肌肉")
    val_idx = get_column_index(header, "MVC_极大值")

    if any(i == -1 for i in [subj_idx, chan_idx, val_idx]):
        print("  错误：MVC_values.csv 缺少必要列")
        sys.exit(1)

    # 修改 zhangyining 的 MVC 值
    updated_count = 0
    old_values = {}
    for row in data_rows:
        if row[subj_idx].strip() == SUBJECT and row[chan_idx].strip() in final_mvc:
            col_name = row[chan_idx].strip()
            old_val = float(row[val_idx])
            new_val = final_mvc[col_name]
            old_values[col_name] = old_val

    # 打印对比
    print(f"\n  {'通道':<30} {'原MVC值':<12} {'新MVC值':<12} {'变化':<10}")
    print("  " + "-" * 66)
    for col_name in CHANNEL_COLS:
        if col_name in old_values and col_name in final_mvc:
            old_val = old_values[col_name]
            new_val = final_mvc[col_name]
            change = new_val - old_val
            change_str = f"{change:+.4f}"
            print(f"  {col_name:<30} {old_val:<12.4f} {new_val:<12.4f} {change_str:<10}")

    # 执行更新
    for row in data_rows:
        if row[subj_idx].strip() == SUBJECT and row[chan_idx].strip() in final_mvc:
            col_name = row[chan_idx].strip()
            row[val_idx] = f"{final_mvc[col_name]:.6f}"
            updated_count += 1

    # 保存
    with open(MVC_CSV_PATH, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(data_rows)

    print(f"\n  ✅ 已更新 {updated_count} 条 MVC 记录")
    print(f"  结果已保存至: {MVC_CSV_PATH}")

    # 4. 清理旧的 Excel 中 zhangyining 的 sheet
    print(f"\n  清理旧的 Excel 分析结果...")
    if os.path.exists(OUTPUT_EXCEL):
        from openpyxl import load_workbook
        wb = load_workbook(OUTPUT_EXCEL)
        removed = False
        if SUBJECT in wb.sheetnames:
            del wb[SUBJECT]
            removed = True
            print(f"  已删除 Sheet: {SUBJECT}")
        if "汇总" in wb.sheetnames:
            del wb["汇总"]
            removed = True
            print(f"  已删除 Sheet: 汇总")
        if removed:
            wb.save(OUTPUT_EXCEL)
            print(f"  已保存更新后的 Excel")
        else:
            print(f"  Excel 中无需要清理的 Sheet")
    else:
        print(f"  Excel 文件不存在，无需清理")

    # 5. 清理旧的 zhangyining 图表
    print(f"\n  清理旧的 zhangyining 图表...")
    if os.path.exists(FIGURES_DIR):
        removed_files = 0
        for fname in os.listdir(FIGURES_DIR):
            if SUBJECT in fname:
                os.remove(os.path.join(FIGURES_DIR, fname))
                removed_files += 1
                print(f"  已删除: {fname}")
        if removed_files == 0:
            print(f"  无相关图表需要清理")
    else:
        print(f"  figures 目录不存在，无需清理")

    print(f"\n✅ MVC 合并完成！请继续运行步骤3和步骤4。")


if __name__ == "__main__":
    main()