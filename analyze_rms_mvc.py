import numpy as np
import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ========== 常量 ==========

CHANNEL_COLS = [
    "Channel 1_三角肌前束",
    "Channel 2_胸锁乳突肌",
    "Channel 3_斜方肌",
    "Channel 4_竖脊肌",
]

FS = 1000                # 采样率 (Hz)
WINDOW_DURATION = 1.0    # RMS 滑动窗口时长 (秒)
WINDOW_SIZE = int(WINDOW_DURATION * FS)

T_START = 10.0    # 分析起始时间 (秒)，跳过前 10s 起始段
T_END   = 160.0   # 分析结束时间 (秒)，取到 160s

PROCESSED_ROOT = r"e:\Coding_projects\SEMG_ANSYS\data_processed"
MVC_CSV_PATH = r"e:\Coding_projects\SEMG_ANSYS\MVC_values.csv"
OUTPUT_EXCEL = r"e:\Coding_projects\SEMG_ANSYS\analysis_results.xlsx"


# ========== 工具函数 ==========

def read_csv_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return None, None
    return rows[0], rows[1:]


def get_column_index(header, col_name):
    for i, h in enumerate(header):
        if h.strip() == col_name:
            return i
    return -1


def load_mvc_values(mvc_path):
    """
    读取 MVC_values.csv，返回字典:
        mvc_lookup[subject][channel_col] = mvc_value
    """
    header, data_rows = read_csv_file(mvc_path)
    if header is None:
        raise FileNotFoundError(f"无法读取 MVC 文件: {mvc_path}")

    # 找到各列的索引
    subj_idx = get_column_index(header, "受试者")
    chan_idx = get_column_index(header, "通道_肌肉")
    val_idx = get_column_index(header, "MVC_极大值")

    if any(i == -1 for i in [subj_idx, chan_idx, val_idx]):
        raise ValueError("MVC_values.csv 缺少必要列")

    lookup = {}
    for row in data_rows:
        subject = row[subj_idx].strip()
        channel_col = row[chan_idx].strip()
        mvc_val = float(row[val_idx])

        if subject not in lookup:
            lookup[subject] = {}
        lookup[subject][channel_col] = mvc_val

    return lookup


def sliding_rms(signal, window_size):
    """
    计算滑动均方根值 (Sliding RMS)
    signal      : 已整流的一维信号
    window_size : 窗口内的数据点数
    return      : RMS 序列 (长度 = len(signal) - window_size + 1)
    """
    squared = signal ** 2
    window = np.ones(window_size) / window_size
    mean_squared = np.convolve(squared, window, mode='valid')
    return np.sqrt(mean_squared)


def compute_percent_mvc_stats(rms_seq, mvc_value):
    """
    将 RMS 序列归一化为 %MVC，并计算 4 个统计指标
    rms_seq   : 滑动 RMS 序列
    mvc_value : 该通道的 MVC 极大值
    return    : (mean, apdf10, apdf50, apdf90)
    """
    if mvc_value <= 0:
        return (0.0, 0.0, 0.0, 0.0)

    percent_mvc = (rms_seq / mvc_value) * 100.0

    mean_val = np.mean(percent_mvc)
    apdf10 = np.percentile(percent_mvc, 10)
    apdf50 = np.percentile(percent_mvc, 50)
    apdf90 = np.percentile(percent_mvc, 90)

    return (mean_val, apdf10, apdf50, apdf90)


# ========== 主程序 ==========

def main():
    # 1. 读取 MVC 参考值
    print("加载 MVC 参考值...")
    mvc_lookup = load_mvc_values(MVC_CSV_PATH)
    print(f"  共加载 {sum(len(v) for v in mvc_lookup.values())} 条 MVC 记录\n")

    # 2. 准备 Excel 工作簿（追加模式：已有则加载，保留旧受试者 Sheet）
    if os.path.exists(OUTPUT_EXCEL):
        wb = load_workbook(OUTPUT_EXCEL)
        # 删除旧的"汇总" Sheet（稍后重新生成）
        if "汇总" in wb.sheetnames:
            del wb["汇总"]
    else:
        wb = Workbook()
        # 删除默认的 "Sheet"
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # 用于收集所有结果（生成汇总 Sheet）
    all_results = []

    # 检查已有 Excel 中已处理过的受试者（从加载的 wb 中读取）
    existing_subjects = set()
    for sn in wb.sheetnames:
        if sn == "汇总":
            continue
        existing_subjects.add(sn)
        ws_existing = wb[sn]
        for row in ws_existing.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                all_results.append((sn, *row))
    if existing_subjects:
        print(f"  已有 Excel 中包含受试者: {existing_subjects}")

    # 3. 遍历每个受试者
    for subject_dir in sorted(os.listdir(PROCESSED_ROOT)):
        subject_path = os.path.join(PROCESSED_ROOT, subject_dir)
        if not os.path.isdir(subject_path):
            continue

        # === 增量处理：已有记录则跳过 ===
        if subject_dir in existing_subjects:
            print(f"  ⏭ 已存在，跳过: {subject_dir}")
            continue

        print(f"=== 处理受试者: {subject_dir} ===")

        # 检查该受试者是否有 MVC 数据
        if subject_dir not in mvc_lookup:
            print(f"  警告：未找到 {subject_dir} 的 MVC 数据，跳过")
            continue

        subject_mvc = mvc_lookup[subject_dir]

        # 筛选"装饰"和"打磨"任务文件
        task_files = []
        for filename in sorted(os.listdir(subject_path)):
            if "最大收缩" in filename:
                continue
            if "装饰" in filename or "打磨" in filename:
                task_files.append(filename)

        if not task_files:
            print(f"  未找到装饰/打磨任务文件")
            continue

        # 为该受试者创建 Sheet
        ws = wb.create_sheet(title=subject_dir)

        # 写表头
        headers = ["动作", "通道_肌肉", "Mean_%MVC", "APDF_10%", "APDF_50%", "APDF_90%"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row_idx = 2

        for task_file in task_files:
            file_path = os.path.join(subject_path, task_file)
            header_row, data_rows = read_csv_file(file_path)
            if header_row is None:
                print(f"  跳过空文件: {task_file}")
                continue

            print(f"  处理: {task_file}")

            for channel_col in CHANNEL_COLS:
                # 获取该通道 MVC 值
                if channel_col not in subject_mvc:
                    print(f"    跳过 {channel_col}（无 MVC 值）")
                    continue
                mvc_val = subject_mvc[channel_col]

                # 提取信号
                col_idx = get_column_index(header_row, channel_col)
                if col_idx == -1:
                    print(f"    文件缺少列: {channel_col}")
                    continue

                signal = np.array(
                    [float(r[col_idx]) for r in data_rows], dtype=np.float64
                )

                # 截取 10s–160s 时间范围
                start_idx = int(T_START * FS)
                end_idx = int(T_END * FS)
                if end_idx > len(signal):
                    print(f"    信号长度 ({len(signal)} 点 = {len(signal)/FS:.1f}s) 不足 {T_END}s，跳过")
                    continue
                signal = signal[start_idx:end_idx]

                # 检查信号长度
                if len(signal) < WINDOW_SIZE:
                    print(f"    信号长度 ({len(signal)}) < 窗口 ({WINDOW_SIZE})，跳过")
                    continue

                # 滑动 RMS
                rms_seq = sliding_rms(signal, WINDOW_SIZE)

                # %MVC 统计
                mean_val, apdf10, apdf50, apdf90 = compute_percent_mvc_stats(
                    rms_seq, mvc_val
                )

                # 写入 Excel
                ws.cell(row=row_idx, column=1, value=task_file).border = thin_border
                ws.cell(row=row_idx, column=2, value=channel_col).border = thin_border
                ws.cell(row=row_idx, column=3, value=round(mean_val, 4)).border = thin_border
                ws.cell(row=row_idx, column=4, value=round(apdf10, 4)).border = thin_border
                ws.cell(row=row_idx, column=5, value=round(apdf50, 4)).border = thin_border
                ws.cell(row=row_idx, column=6, value=round(apdf90, 4)).border = thin_border

                # 收集汇总数据
                all_results.append(
                    (subject_dir, task_file, channel_col, mean_val, apdf10, apdf50, apdf90)
                )

                row_idx += 1

        # 自动调整列宽
        for col in range(1, 7):
            max_len = len(str(ws.cell(row=1, column=col).value))
            for r in range(2, row_idx):
                cell_len = len(str(ws.cell(row=r, column=col).value or ""))
                max_len = max(max_len, cell_len)
            ws.column_dimensions[chr(64 + col)].width = max_len + 4

    # 4. 创建汇总 Sheet
    if all_results:
        ws_summary = wb.create_sheet(title="汇总", index=0)
        summary_headers = ["受试者", "动作", "通道_肌肉", "Mean_%MVC", "APDF_10%", "APDF_50%", "APDF_90%"]
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for r_idx, (subj, task, ch, mean_v, p10, p50, p90) in enumerate(all_results, 2):
            ws_summary.cell(row=r_idx, column=1, value=subj).border = thin_border
            ws_summary.cell(row=r_idx, column=2, value=task).border = thin_border
            ws_summary.cell(row=r_idx, column=3, value=ch).border = thin_border
            ws_summary.cell(row=r_idx, column=4, value=round(mean_v, 4)).border = thin_border
            ws_summary.cell(row=r_idx, column=5, value=round(p10, 4)).border = thin_border
            ws_summary.cell(row=r_idx, column=6, value=round(p50, 4)).border = thin_border
            ws_summary.cell(row=r_idx, column=7, value=round(p90, 4)).border = thin_border

        for col in range(1, 8):
            ws_summary.column_dimensions[chr(64 + col)].width = 18

    # 5. 删除残留的默认空 Sheet（如有）
    for sn in wb.sheetnames:
        if sn == "Sheet":
            wb.remove(wb[sn])
            break

    # 6. 保存
    wb.save(OUTPUT_EXCEL)
    print(f"\n✅ 分析完成！共处理 {len(all_results)} 条记录")
    print(f"   结果已保存至: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()